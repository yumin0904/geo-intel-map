"""
scripts/load_external_data.py

IA-Engine-B1/Cycle-6A: SIPRI · COW · Kiel · EIA · CSIS · V-DEM · SIPRI Arms · COW Wars 외부 정형 데이터를 intel.db에 적재한다.

기본 동작: data/external/*_seed.csv 시드 파일 사용 (오프라인 작동)
업데이트 모드 (--update): 원본 사이트에서 최신 Excel/ZIP 다운로드 시도

실행:
    python3 backend/scripts/load_external_data.py          # 시드 데이터 적재
    python3 backend/scripts/load_external_data.py --update # 최신 데이터 다운로드 후 적재
    python3 backend/scripts/load_external_data.py --source sipri  # 소스 지정
"""
from __future__ import annotations

import argparse
import csv
import io
import logging
import os
import sqlite3
import sys
import zipfile
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

_ROOT    = Path(__file__).resolve().parents[2]
_DB_PATH = _ROOT / "backend" / "db" / "intel.db"
_EXT_DIR = _ROOT / "data" / "external"


# ── DB 초기화 (테이블이 없으면 생성) ──────────────────────────────────────────

def _ensure_tables(con: sqlite3.Connection) -> None:
    con.executescript("""
    CREATE TABLE IF NOT EXISTS sipri_milex (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        iso3 TEXT NOT NULL, country_name TEXT,
        year INTEGER NOT NULL,
        gdp_pct REAL, usd_mn_2022 REAL,
        UNIQUE(iso3, year)
    );
    CREATE TABLE IF NOT EXISTS cow_alliances (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        iso3_a TEXT NOT NULL, iso3_b TEXT NOT NULL,
        name_a TEXT, name_b TEXT,
        start_year INTEGER, end_year INTEGER,
        alliance_type TEXT,
        UNIQUE(iso3_a, iso3_b, start_year)
    );
    CREATE TABLE IF NOT EXISTS kiel_ukraine_support (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donor_iso3 TEXT, donor_name TEXT NOT NULL,
        military_eur_bn REAL DEFAULT 0,
        financial_eur_bn REAL DEFAULT 0,
        humanitarian_eur_bn REAL DEFAULT 0,
        total_eur_bn REAL DEFAULT 0,
        data_period TEXT,
        updated_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ', 'now')),
        UNIQUE(donor_iso3, data_period)
    );
    CREATE TABLE IF NOT EXISTS sipri_arms_transfers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        supplier_iso3 TEXT NOT NULL, supplier_name TEXT,
        recipient_iso3 TEXT NOT NULL, recipient_name TEXT,
        year INTEGER NOT NULL,
        tiv_mn REAL, weapon_category TEXT, notes TEXT,
        UNIQUE(supplier_iso3, recipient_iso3, year, weapon_category)
    );
    CREATE TABLE IF NOT EXISTS vdem_index (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        iso3 TEXT NOT NULL, country_name TEXT,
        year INTEGER NOT NULL,
        v2x_libdem REAL, v2x_regime INTEGER,
        v2x_polyarchy REAL, v2x_corr REAL, notes TEXT,
        UNIQUE(iso3, year)
    );
    CREATE TABLE IF NOT EXISTS cow_wars (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        war_id INTEGER UNIQUE, war_name TEXT NOT NULL,
        start_year INTEGER, end_year INTEGER,
        side_a_iso3 TEXT, side_b_iso3 TEXT,
        region TEXT, battle_deaths INTEGER,
        outcome INTEGER, relevance_tag TEXT
    );
    CREATE TABLE IF NOT EXISTS eia_energy (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        iso3 TEXT NOT NULL, country_name TEXT,
        crude_prod_mbpd REAL, natgas_prod_bcfd REAL, oil_export_mbpd REAL,
        data_year INTEGER,
        UNIQUE(iso3, data_year)
    );
    CREATE TABLE IF NOT EXISTS csis_cyber_incidents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        incident_id TEXT UNIQUE,
        incident_date TEXT, actor_iso3 TEXT, actor_group TEXT,
        victim_iso3 TEXT, victim_sector TEXT, incident_type TEXT,
        title TEXT, description TEXT
    );
    """)
    con.commit()


# ── SIPRI Military Expenditure ─────────────────────────────────────────────────

def _load_sipri_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 SIPRI 데이터 적재."""
    seed = _EXT_DIR / "sipri_milex_seed.csv"
    if not seed.exists():
        logger.warning("SIPRI 시드 파일 없음: %s", seed)
        return 0

    rows = 0
    with open(seed, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(",")
            if parts[0] == "iso3":
                continue  # 헤더
            iso3, name, year = parts[0], parts[1], int(parts[2])
            gdp  = float(parts[3]) if parts[3] else None
            usd  = float(parts[4]) if parts[4] else None
            con.execute(
                "INSERT OR REPLACE INTO sipri_milex (iso3, country_name, year, gdp_pct, usd_mn_2022)"
                " VALUES (?,?,?,?,?)",
                (iso3, name, year, gdp, usd),
            )
            rows += 1
    con.commit()
    return rows


def _load_sipri_excel(con: sqlite3.Connection, path: Path) -> int:
    """SIPRI Excel 파일에서 'Share of GDP' 시트 파싱."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl 미설치 — pip install openpyxl")
        return 0

    wb   = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # 시트명은 버전마다 다를 수 있음
    sheet_name = next(
        (s for s in wb.sheetnames if "GDP" in s.upper() or "gdp" in s.lower()),
        None,
    )
    if not sheet_name:
        logger.error("SIPRI Excel에서 GDP 시트를 찾을 수 없습니다. 시트 목록: %s", wb.sheetnames)
        return 0

    ws   = wb[sheet_name]
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return 0

    # 헤더 행에서 연도 위치 찾기
    header = rows_data[0]
    year_cols: list[tuple[int, int]] = []  # (col_idx, year)
    for i, v in enumerate(header):
        try:
            y = int(v)
            if 2000 <= y <= 2030:
                year_cols.append((i, y))
        except (TypeError, ValueError):
            pass

    rows_inserted = 0
    for row in rows_data[1:]:
        iso3 = str(row[1]).strip() if row[1] else None
        name = str(row[0]).strip() if row[0] else None
        if not iso3 or len(iso3) != 3:
            continue
        for col_idx, year in year_cols:
            val = row[col_idx]
            if val is None:
                continue
            try:
                gdp_pct = float(val)
            except (TypeError, ValueError):
                continue
            con.execute(
                "INSERT OR REPLACE INTO sipri_milex (iso3, country_name, year, gdp_pct)"
                " VALUES (?,?,?,?)",
                (iso3, name, year, gdp_pct),
            )
            rows_inserted += 1

    con.commit()
    wb.close()
    return rows_inserted


def _download_sipri() -> Path | None:
    """SIPRI Excel 최신 파일 다운로드 시도."""
    try:
        import requests  # type: ignore
    except ImportError:
        logger.error("requests 미설치")
        return None

    # SIPRI는 고정 URL이 없으므로 최근 알려진 경로 시도
    urls = [
        "https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2023.xlsx",
        "https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2022.xlsx",
    ]
    dest = _EXT_DIR / "sipri_milex_latest.xlsx"
    for url in urls:
        try:
            r = requests.get(url, timeout=30)
            if r.status_code == 200:
                dest.write_bytes(r.content)
                logger.info("SIPRI 다운로드 완료: %s", dest)
                return dest
        except Exception as e:
            logger.debug("SIPRI URL 실패 %s: %s", url, e)
    logger.warning("SIPRI 자동 다운로드 실패 — 시드 데이터 사용")
    return None


# ── COW Formal Alliances ───────────────────────────────────────────────────────

def _load_cow_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 COW 동맹 데이터 적재."""
    seed = _EXT_DIR / "cow_alliances_seed.csv"
    if not seed.exists():
        logger.warning("COW 시드 파일 없음: %s", seed)
        return 0

    rows = 0
    with open(seed, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(",")
            if parts[0] == "iso3_a":
                continue
            iso3_a, iso3_b = parts[0], parts[1]
            name_a, name_b = parts[2], parts[3]
            start  = int(parts[4]) if parts[4] else None
            end    = int(parts[5]) if parts[5] else None
            atype  = parts[6] if len(parts) > 6 else "defense"

            # 중복 처리 — 프랑스처럼 ended 버전과 active 버전이 겹치는 경우 skip
            try:
                con.execute(
                    "INSERT OR IGNORE INTO cow_alliances"
                    " (iso3_a, iso3_b, name_a, name_b, start_year, end_year, alliance_type)"
                    " VALUES (?,?,?,?,?,?,?)",
                    (iso3_a, iso3_b, name_a, name_b, start, end, atype),
                )
                rows += 1
            except sqlite3.IntegrityError:
                pass

    con.commit()
    return rows


def _download_cow() -> Path | None:
    """COW v4.1 ZIP 다운로드."""
    try:
        import requests  # type: ignore
    except ImportError:
        return None

    url  = "https://correlatesofwar.org/wp-content/uploads/alliance_v4.1_by_directed_yearly.zip"
    dest = _EXT_DIR / "cow_alliances_v4.1.zip"
    try:
        r = requests.get(url, timeout=60)
        if r.status_code == 200:
            dest.write_bytes(r.content)
            logger.info("COW 다운로드 완료: %s", dest)
            return dest
    except Exception as e:
        logger.warning("COW 다운로드 실패: %s", e)
    return None


def _load_cow_zip(con: sqlite3.Connection, path: Path) -> int:
    """COW ZIP 안의 CSV 파싱 (directed dyad yearly 형식)."""
    # COW ISO3 ↔ ccode 매핑 (주요국)
    _CCODE_ISO3 = {
        2: "USA", 200: "GBR", 220: "FRA", 255: "DEU", 325: "ITA",
        365: "RUS", 710: "CHN", 740: "JPN", 732: "KOR", 731: "PRK",
        630: "IRN", 696: "ISR", 663: "ISR", 670: "SAU", 620: "LBY",
        816: "VNM", 820: "MYS", 840: "PHL", 900: "AUS", 920: "NZL",
        750: "IND", 770: "PAK", 780: "LKA", 640: "TUR", 360: "POL",
        375: "FIN", 380: "SWE", 385: "NOR", 390: "DNK", 310: "HUN",
        315: "CZE", 290: "BGR", 367: "EST", 368: "LVA", 369: "LTU",
        230: "ESP", 235: "PRT", 305: "AUT", 211: "BEL", 210: "NLD",
        355: "BGR", 349: "ALB", 344: "HRV", 349: "MNE", 343: "MKD",
        350: "SVN", 317: "SVK", 316: "ROU", 339: "ALB",
        100: "COL", 130: "ECU", 140: "BRA", 160: "ARG",
        350: "GRC", 352: "CYP", 395: "ISL",
        600: "MAR", 615: "ALG", 645: "IRQ", 652: "EGY", 660: "LBN",
        666: "ISR", 678: "YEM", 694: "QAT", 692: "BHR", 690: "SAU",
        698: "OMN", 694: "ARE", 703: "AFG", 704: "PAK", 705: "IND",
        712: "MNG", 713: "TWN", 800: "THA", 811: "KHM", 812: "LAO",
        817: "VNM", 820: "MYS", 830: "SGP", 835: "BRN", 850: "IDN",
        900: "AUS", 920: "NZL", 950: "FJI",
        230: "ESP", 232: "AND",
        550: "ETH", 560: "ZAF", 571: "ZWE", 572: "ZMB",
    }
    _ATYPE_MAP = {
        1: "defense", 2: "neutrality", 3: "nonaggression", 4: "consultation",
    }
    try:
        with zipfile.ZipFile(path) as zf:
            csv_name = next(n for n in zf.namelist() if n.endswith(".csv"))
            with zf.open(csv_name) as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                rows_inserted = 0
                for row in reader:
                    try:
                        c1 = int(row.get("ccode1", 0))
                        c2 = int(row.get("ccode2", 0))
                        iso_a = _CCODE_ISO3.get(c1)
                        iso_b = _CCODE_ISO3.get(c2)
                        if not iso_a or not iso_b:
                            continue
                        sy = int(row.get("dyad_st_year", 0) or 0)
                        ey_raw = row.get("dyad_end_year", "")
                        ey = int(ey_raw) if ey_raw and ey_raw.strip() else None
                        # 종료된 동맹 중 1990년 이전 종료는 제외
                        if ey and ey < 1990:
                            continue
                        atype_code = int(row.get("ss_type1", 0) or 0)
                        atype = _ATYPE_MAP.get(atype_code, "unknown")
                        con.execute(
                            "INSERT OR IGNORE INTO cow_alliances"
                            " (iso3_a, iso3_b, start_year, end_year, alliance_type)"
                            " VALUES (?,?,?,?,?)",
                            (iso_a, iso_b, sy, ey, atype),
                        )
                        rows_inserted += 1
                    except (ValueError, KeyError):
                        continue
                con.commit()
                return rows_inserted
    except Exception as e:
        logger.error("COW ZIP 파싱 실패: %s", e)
        return 0


# ── Kiel Ukraine Support Tracker ──────────────────────────────────────────────

def _load_kiel_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 Kiel 데이터 적재."""
    seed = _EXT_DIR / "kiel_ukraine_support_seed.csv"
    if not seed.exists():
        logger.warning("Kiel 시드 파일 없음: %s", seed)
        return 0

    rows = 0
    with open(seed, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(",")
            if parts[0] == "donor_iso3":
                continue
            donor_iso3  = parts[0] or None
            donor_name  = parts[1]
            mil   = float(parts[2]) if parts[2] else 0.0
            fin   = float(parts[3]) if parts[3] else 0.0
            hum   = float(parts[4]) if parts[4] else 0.0
            total = float(parts[5]) if parts[5] else 0.0
            period = parts[6] if len(parts) > 6 else "2022-01~2024-06"
            con.execute(
                "INSERT OR REPLACE INTO kiel_ukraine_support"
                " (donor_iso3, donor_name, military_eur_bn, financial_eur_bn,"
                "  humanitarian_eur_bn, total_eur_bn, data_period)"
                " VALUES (?,?,?,?,?,?,?)",
                (donor_iso3, donor_name, mil, fin, hum, total, period),
            )
            rows += 1
    con.commit()
    return rows


# ── EIA Energy ────────────────────────────────────────────────────────────────

def _load_eia_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 EIA 에너지 데이터 적재."""
    seed = _EXT_DIR / "eia_energy_seed.csv"
    if not seed.exists():
        logger.warning("EIA 시드 파일 없음: %s", seed)
        return 0

    rows = 0
    with open(seed, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(",")
            if parts[0] == "iso3":
                continue
            iso3 = parts[0]
            name = parts[1]
            crude  = float(parts[2]) if parts[2] else None
            natgas = float(parts[3]) if parts[3] else None
            export = float(parts[4]) if parts[4] else None
            year   = int(parts[5]) if parts[5] else 2023
            con.execute(
                "INSERT OR REPLACE INTO eia_energy"
                " (iso3, country_name, crude_prod_mbpd, natgas_prod_bcfd, oil_export_mbpd, data_year)"
                " VALUES (?,?,?,?,?,?)",
                (iso3, name, crude, natgas, export, year),
            )
            rows += 1
    con.commit()
    return rows


def _download_eia(api_key: str) -> int:
    """EIA API v2로 주요국 원유 생산량 최신 데이터 다운로드."""
    try:
        import requests  # type: ignore
    except ImportError:
        return 0

    # EIA API v2: 국제 원유 생산량
    url = "https://api.eia.gov/v2/international/data/"
    params = {
        "api_key": api_key,
        "frequency": "annual",
        "data[0]": "value",
        "facets[activityId][]": "1",   # production
        "facets[productId][]": "57",   # crude oil
        "facets[unit][]": "TBPD",
        "sort[0][column]": "period",
        "sort[0][direction]": "desc",
        "length": 500,
        "offset": 0,
    }
    try:
        r = requests.get(url, params=params, timeout=30)
        if r.status_code != 200:
            logger.warning("EIA API %d", r.status_code)
            return 0
        data = r.json().get("response", {}).get("data", [])
        logger.info("EIA API 수신: %d건", len(data))
        return len(data)
    except Exception as e:
        logger.warning("EIA API 실패: %s", e)
        return 0


# ── CSIS Cyber Incidents ───────────────────────────────────────────────────────

def _load_csis_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 CSIS 사이버 사건 데이터 적재."""
    seed = _EXT_DIR / "csis_cyber_seed.csv"
    if not seed.exists():
        logger.warning("CSIS 시드 파일 없음: %s", seed)
        return 0

    rows = 0
    with open(seed, encoding="utf-8") as f:
        reader = csv.DictReader(
            (line for line in f if not line.startswith("#")),
        )
        for row in reader:
            try:
                con.execute(
                    "INSERT OR IGNORE INTO csis_cyber_incidents"
                    " (incident_id, incident_date, actor_iso3, actor_group,"
                    "  victim_iso3, victim_sector, incident_type, title, description)"
                    " VALUES (?,?,?,?,?,?,?,?,?)",
                    (
                        row.get("incident_id"),
                        row.get("date"),
                        row.get("actor_iso3") or None,
                        row.get("actor_group") or None,
                        row.get("victim_iso3") or None,
                        row.get("victim_sector") or None,
                        row.get("incident_type") or None,
                        row.get("title"),
                        row.get("description"),
                    ),
                )
                rows += 1
            except (sqlite3.IntegrityError, KeyError):
                pass
    con.commit()
    return rows


def _load_sipri_arms_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 SIPRI Arms Transfers 데이터 적재."""
    seed = _EXT_DIR / "sipri_arms_seed.csv"
    if not seed.exists():
        logger.warning("SIPRI Arms 시드 파일 없음: %s", seed)
        return 0
    rows = 0
    with open(seed, encoding="utf-8") as f:
        reader = csv.DictReader(line for line in f if not line.startswith("#"))
        for row in reader:
            try:
                con.execute(
                    "INSERT OR IGNORE INTO sipri_arms_transfers"
                    " (supplier_iso3, supplier_name, recipient_iso3, recipient_name,"
                    "  year, tiv_mn, weapon_category, notes)"
                    " VALUES (?,?,?,?,?,?,?,?)",
                    (
                        row["supplier_iso3"], row.get("supplier_name"),
                        row["recipient_iso3"], row.get("recipient_name"),
                        int(row["year"]),
                        float(row["tiv_mn"]) if row.get("tiv_mn") else None,
                        row.get("weapon_category"), row.get("notes"),
                    ),
                )
                rows += 1
            except (sqlite3.IntegrityError, KeyError, ValueError):
                pass
    con.commit()
    return rows


def _load_vdem_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 V-DEM 민주주의 지수 데이터 적재."""
    seed = _EXT_DIR / "vdem_seed.csv"
    if not seed.exists():
        logger.warning("V-DEM 시드 파일 없음: %s", seed)
        return 0
    rows = 0
    with open(seed, encoding="utf-8") as f:
        reader = csv.DictReader(line for line in f if not line.startswith("#"))
        for row in reader:
            try:
                con.execute(
                    "INSERT OR REPLACE INTO vdem_index"
                    " (iso3, country_name, year, v2x_libdem, v2x_regime,"
                    "  v2x_polyarchy, v2x_corr, notes)"
                    " VALUES (?,?,?,?,?,?,?,?)",
                    (
                        row["iso3"], row.get("country_name"),
                        int(row["year"]),
                        float(row["v2x_libdem"]) if row.get("v2x_libdem") else None,
                        int(row["v2x_regime"]) if row.get("v2x_regime") else None,
                        float(row["v2x_polyarchy"]) if row.get("v2x_polyarchy") else None,
                        float(row["v2x_corr"]) if row.get("v2x_corr") else None,
                        row.get("notes"),
                    ),
                )
                rows += 1
            except (sqlite3.IntegrityError, KeyError, ValueError):
                pass
    con.commit()
    return rows


def _load_cow_wars_seed(con: sqlite3.Connection) -> int:
    """시드 CSV에서 COW Wars 전쟁 데이터 적재."""
    seed = _EXT_DIR / "cow_wars_seed.csv"
    if not seed.exists():
        logger.warning("COW Wars 시드 파일 없음: %s", seed)
        return 0
    rows = 0
    with open(seed, encoding="utf-8") as f:
        reader = csv.DictReader(line for line in f if not line.startswith("#"))
        for row in reader:
            try:
                con.execute(
                    "INSERT OR IGNORE INTO cow_wars"
                    " (war_id, war_name, start_year, end_year,"
                    "  side_a_iso3, side_b_iso3, region,"
                    "  battle_deaths, outcome, relevance_tag)"
                    " VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (
                        int(row["war_id"]) if row.get("war_id") else None,
                        row["war_name"],
                        int(row["start_year"]) if row.get("start_year") else None,
                        int(row["end_year"]) if row.get("end_year") else None,
                        row.get("side_a_iso3"), row.get("side_b_iso3"),
                        row.get("region"),
                        int(row["battle_deaths"]) if row.get("battle_deaths") else None,
                        int(row["outcome"]) if row.get("outcome") else None,
                        row.get("relevance_tag"),
                    ),
                )
                rows += 1
            except (sqlite3.IntegrityError, KeyError, ValueError):
                pass
    con.commit()
    return rows


def _download_csis() -> Path | None:
    """CSIS Significant Cyber Incidents Excel 다운로드 시도."""
    try:
        import requests  # type: ignore
    except ImportError:
        return None

    urls = [
        "https://www.csis.org/programs/strategic-technologies-program/significant-cyber-incidents",
    ]
    # CSIS는 페이지에 Excel 링크가 있지만 직접 URL이 자주 바뀜 → 시드 우선
    logger.info("CSIS 자동 다운로드는 지원되지 않음 — 시드 데이터 사용")
    return None


# ── 메인 ──────────────────────────────────────────────────────────────────────

def main(sources: list[str], update: bool) -> None:
    _EXT_DIR.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(_DB_PATH)
    _ensure_tables(con)

    results: dict[str, int] = {}

    if "sipri" in sources:
        path = None
        if update:
            path = _download_sipri()
        if path and path.exists():
            n = _load_sipri_excel(con, path)
        else:
            n = _load_sipri_seed(con)
        results["sipri"] = n
        logger.info("SIPRI 적재 완료: %d행", n)

    if "cow" in sources:
        path = None
        if update:
            path = _download_cow()
        if path and path.exists():
            n = _load_cow_zip(con, path)
        else:
            n = _load_cow_seed(con)
        results["cow"] = n
        logger.info("COW 적재 완료: %d행", n)

    if "kiel" in sources:
        n = _load_kiel_seed(con)
        results["kiel"] = n
        logger.info("Kiel 적재 완료: %d행", n)

    if "eia" in sources:
        api_key = os.environ.get("EIA_API_KEY", "")
        if update and api_key:
            n = _download_eia(api_key)
            if n == 0:
                n = _load_eia_seed(con)
        else:
            n = _load_eia_seed(con)
        results["eia"] = n
        logger.info("EIA 적재 완료: %d행", n)

    if "csis" in sources:
        n = _load_csis_seed(con)
        results["csis"] = n
        logger.info("CSIS 적재 완료: %d행", n)

    if "sipri_arms" in sources:
        n = _load_sipri_arms_seed(con)
        results["sipri_arms"] = n
        logger.info("SIPRI Arms 적재 완료: %d행", n)

    if "vdem" in sources:
        n = _load_vdem_seed(con)
        results["vdem"] = n
        logger.info("V-DEM 적재 완료: %d행", n)

    if "cow_wars" in sources:
        n = _load_cow_wars_seed(con)
        results["cow_wars"] = n
        logger.info("COW Wars 적재 완료: %d행", n)

    con.close()

    total = sum(results.values())
    logger.info("=== 전체 %d행 적재 완료 ===", total)
    for src, n in results.items():
        logger.info("  %-8s %d행", src, n)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="외부 정형 데이터 적재")
    parser.add_argument("--source", default="all",
                        help="all | sipri | cow | kiel | eia | csis | sipri_arms | vdem | cow_wars (쉼표로 복수 지정)")
    parser.add_argument("--update", action="store_true",
                        help="원본 사이트에서 최신 데이터 다운로드 시도")
    args = parser.parse_args()

    srcs = ["sipri", "cow", "kiel", "eia", "csis", "sipri_arms", "vdem", "cow_wars"] \
        if args.source == "all" \
        else [s.strip() for s in args.source.split(",")]
    main(srcs, args.update)
